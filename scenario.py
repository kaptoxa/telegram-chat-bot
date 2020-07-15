import exceptions
import os
from shutil import copyfile

from openpyxl import load_workbook
from pprint import pprint
from itertools import groupby


SCENARIO_PATH = os.getenv('SCENARIO_PATH')


class Scenario():

    def read_from_file(self, filename):
        # читаем из файла сценарий, на выход словар
        # словарь с ключём - номер фразы бота + значения (ссылки и ответы)
        wb = load_workbook(filename)
        ws = wb.active

        phrases = {}
        backs = {}

        ws.delete_rows(1)

        p_indexes = [(cell.value, cell.row) for cell in ws['A'] if cell.value]
        if len(p_indexes) < 2:
            raise exceptions.NotCorrectScenarioFile('У вас слишком короткий сценарий :)')
        for key, row in p_indexes:
            text = ws.cell(row=row, column=2).value
            if not text:
                raise exceptions.NotCorrectScenarioFile(f"Нет текста для фразы {key}")
            phrases[key] = {'text': text}

        a_indexes = [(cell.value, cell.row) for cell in ws['C'] if cell.value]
        answers = [(text, row, ws.cell(row=row, column=4).value) for text, row in a_indexes]  # в 4 столбце ссылка на фразу

        def index_phrase_for_answer(row):
            return max(filter(lambda pi: pi[1] <= row[1], p_indexes), key=lambda pi: pi[1])[0]

        for key, group in groupby(answers, index_phrase_for_answer):
            phrases[key]['answers'] = []
            for text, row, link in group:
                if not link:
                    raise exceptions.NotCorrectScenarioFile(f"Нет ссылки для ответа в строке {row+1}")
                    return None
                phrases[key]['answers'].append((text, link))
                backs[link] = key

        for key in phrases.keys():
            phrases[key]['back'] = backs.get(key, 1)

        wb.close()
        return phrases

    def load_checked_file(self):
        try:
            self.scenario = self.read_from_file(SCENARIO_PATH)
        except Exception as e:
            print(e)
            return 'Scenario cannot be parsed!'
        return 'Scenario successfully parsed.'

    def __init__(self):
        print(self.load_checked_file())

    def check(self, phrases):
        pprint(phrases)
        for key in phrases.keys():
            if 'answers' in phrases[key]:
                for answer, link in phrases[key]['answers']:
                    if link not in phrases:
                        raise exceptions.NotCorrectScenarioLinks(f"Link to unknown phrase - {link}")
                        return False
        return True

    def load(self, filename):
        phrases = self.read_from_file(filename)
        print('Сценарий распаршен!')
        if self.check(phrases):
            copyfile(filename, 'scenario.xlsx')
            print('Заменяем файл сценария.')
            self.scenario = phrases
            return True
        return False

    def get(self, state):
        return self.scenario[state]

    def show(self):
        pprint(self.scenario)

    def is_leaf(self, state):
         return 'answers' not in self.scenario[state] or len(self.scenario[state]['answers']) == 0


scenario = Scenario()
